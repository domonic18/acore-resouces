from __future__ import annotations

from pathlib import Path
from typing import Any, cast

import yaml
from openpyxl import load_workbook

from app.core.config import settings
from app.schemas.resource import Mount, Npc, Pet, Resource

SCHEMA_MAP = {
    "mount": Mount,
    "pet": Pet,
    "npc": Npc,
}


def _load_mapping(resource_type: str) -> dict[str, Any]:
    mapping_path = settings.mapping_dir / f"{resource_type}.mapping.yaml"
    with mapping_path.open("r", encoding="utf-8") as f:
        return cast(dict[str, Any], yaml.safe_load(f))


def _set_nested(data: dict[str, Any], key: str, value: Any) -> None:
    parts = key.split(".")
    current = data
    for part in parts[:-1]:
        if part not in current or not isinstance(current[part], dict):
            current[part] = {}
        current = current[part]
    current[parts[-1]] = value


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().upper() in ("Y", "YES", "TRUE", "1", "是")


def _parse_number(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    try:
        text = str(value).strip()
        if not text:
            return None
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return None


NUMERIC_FIELD_KEYS = {
    "id",
    "drop.entry",
    "drop.rate",
    "dbc.creature_model_data.id",
    "dbc.creature_model_data.flags",
    "dbc.creature_display_info.id",
    "dbc.creature_display_info.model_id",
    "dbc.creature_display_info.sound_id",
    "dbc.creature_display_info.extra_display_information_id",
    "dbc.creature_display_info.scale",
    "dbc.creature_display_info.opacity",
    "db.creature_template.entry",
    "db.creature_template.modelid1",
    "db.creature_template.faction",
    "db.creature_template.type",
    "db.creature_template.unit_class",
    "db.creature_template.unit_flag2",
    "db.creature_model_info.display_id",
    "db.creature_template.bounding_radius",
    "db.creature_template.combat_reach",
    "db.creature_template.gender",
    "db.creature_template.display_id_other_gender",
    "dbc.spell.id",
    "dbc.spell.visual_id",
    "dbc.spell.icon_id",
    "db.item_template.entry",
    "db.item_template.displayid",
    "db.item_template.Quality",
    "db.item_template.AllowableClass",
    "db.item_template.AllowableRace",
    "db.item_template.spellid_2",
    "dbc.item.id",
    "dbc.item.class",
    "dbc.item.subclass",
    "dbc.item.material",
    "dbc.item.quality",
    "dbc.item.display_id",
    "dbc.item.inventory_type",
    "dbc.item.sheath",
}


def _row_to_resource(
    row: tuple[Any, ...],
    mapping: dict[str, Any],
) -> Resource:
    resource_type = mapping["resource_type"]
    schema_cls = SCHEMA_MAP[resource_type]
    data: dict[str, Any] = {"resource_type": resource_type}

    for field_key, col_idx in mapping["fields"].items():
        value = row[col_idx] if col_idx < len(row) else None

        if value is None or (isinstance(value, str) and not value.strip()):
            continue

        if field_key in ("debug_passed", "added"):
            value = _parse_bool(value)
        elif field_key in NUMERIC_FIELD_KEYS:
            value = _parse_number(value)
        elif isinstance(value, str):
            value = value.strip()

        _set_nested(data, field_key, value)

    return cast(Resource, schema_cls(**data))


def _save_resource_with_unique_name(
    resource: Resource,
    seen_filenames: set[str],
) -> None:
    from app.services.resource_store import load_resource, save_resource

    existing = load_resource(resource.resource_type, resource.id)
    if existing is not None:
        resource = _merge_resources(existing, resource)

    base_name = f"{resource.id:04d}-{resource.model_folder}"
    if base_name not in seen_filenames:
        seen_filenames.add(base_name)
        save_resource(resource)
        return

    suffix = 2
    while f"{base_name}-{suffix}" in seen_filenames:
        suffix += 1
    seen_filenames.add(f"{base_name}-{suffix}")
    save_resource(resource, filename_suffix=str(suffix))


def _merge_resources(existing: Resource, new: Resource) -> Resource:
    """将新导入的数据合并到已有资源中，保留已有字段 unless 新数据提供了非空值。"""
    schema_cls = SCHEMA_MAP[new.resource_type]
    base = existing.model_dump()
    update = new.model_dump()

    def deep_merge(base_dict: dict[str, Any], update_dict: dict[str, Any]) -> dict[str, Any]:
        result = dict(base_dict)
        for key, value in update_dict.items():
            if value is None or value == {}:
                continue
            if isinstance(value, dict) and isinstance(result.get(key), dict):
                result[key] = deep_merge(result[key], value)
            else:
                result[key] = value
        return result

    merged = deep_merge(base, update)
    return cast(Resource, schema_cls(**merged))


def import_xlsx(
    input_path: Path,
    resource_type: str,
    *,
    dry_run: bool = False,
    limit: int | None = None,
) -> dict[str, Any]:
    mapping = _load_mapping(resource_type)

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[mapping["sheet"]]

    data_start_row = mapping.get("data_start_row", 3)
    created = 0
    updated = 0
    errors: list[dict[str, Any]] = []
    resources: list[Resource] = []
    seen_filenames: set[str] = set()
    last_model_folder: str | None = None

    id_source = mapping.get("id_source", "column")
    id_col = mapping["fields"].get("id")
    required_field = mapping.get("required_field")
    required_col = mapping["fields"].get(required_field) if required_field else None

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row
    ):
        if limit is not None and created >= limit:
            break

        if not row:
            continue

        if all(cell is None for cell in row):
            continue

        if required_col is not None:
            required_value = row[required_col] if required_col < len(row) else None
            if required_value is None or (
                isinstance(required_value, str) and not required_value.strip()
            ):
                continue

        if id_source == "row":
            id_value: int | float | None = row_idx
        elif id_col is not None:
            id_value = _parse_number(row[id_col]) if id_col < len(row) else None
        else:
            id_value = None

        model_folder_col = mapping["fields"].get("model_folder")
        raw_model_folder = (
            str(row[model_folder_col]).strip()
            if model_folder_col is not None
            and model_folder_col < len(row)
            and row[model_folder_col]
            else ""
        )

        if id_value is None:
            # 当前行没有 ID，无法构成独立资源；保留 model_folder 供后续合并单元格行继承
            if raw_model_folder:
                last_model_folder = raw_model_folder
            continue

        model_folder_value = raw_model_folder or last_model_folder
        if not model_folder_value:
            errors.append(
                {
                    "row": row_idx,
                    "error": f"id={id_value} 缺少 model_folder，且无法从上方合并单元格继承",
                }
            )
            continue

        last_model_folder = model_folder_value

        # 将 ID 与继承得到的 model_folder 回填到行数据中，避免 Pydantic 校验失败
        row_list = list(row)
        if id_col is not None and id_col < len(row_list):
            row_list[id_col] = id_value
        if model_folder_col is not None and model_folder_col < len(row_list):
            row_list[model_folder_col] = model_folder_value
        enriched_row = tuple(row_list)

        try:
            resource = _row_to_resource(enriched_row, mapping)
            resources.append(resource)
            if not dry_run:
                _save_resource_with_unique_name(resource, seen_filenames)
            created += 1
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": row_idx, "error": str(exc)})

    workbook.close()

    return {
        "resource_type": resource_type,
        "dry_run": dry_run,
        "created": created,
        "updated": updated,
        "errors": errors,
        "sample": resources[:3] if dry_run else [],
    }
